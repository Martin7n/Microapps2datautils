from datetime import datetime, date
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from data_mask.file_dataclass import DATAMaskrecords
from data_mask.datamask_processor import data_mask
from esg_vehicles.data_collections.headers_fields import HEADER_MAP, OUTPUT_FIELDS
from esg_vehicles.processors.processor import process_records
from excel_read_write.processor_helper import prepare_excel_value
from excel_read_write.settings import DATA_PROCESSING_TYPE


def read_parse_to_class(filename, type_processing):
    wb = load_workbook(filename, data_only=True)
    ws = wb.active

    headers = next(
        ws.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True
        )
    )

    records = []
    data_record = DATA_PROCESSING_TYPE.get(type_processing, DATA_PROCESSING_TYPE["default"])

    for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row,
            values_only=True):

        row_dict = dict(zip(headers, row))

        values = {}

        for excel_name, python_name in HEADER_MAP.items():
            values[python_name] = row_dict.get(excel_name)
        # record = ESGRecord(**values)
        row_dict = dict(zip(headers, row))
        record = data_record(original=row_dict, **values)

        records.append(record)
    # print(records)
    return records




def write_records(filename, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Processed Data"
    a_counter_is_never_too_much = 0
    if not records:
        return

    original_headers = list(records[0].original.keys())
    added_headers = list(OUTPUT_FIELDS.keys())
    ws.append(original_headers + added_headers)


    for record in records:
        a_counter_is_never_too_much +=1
        row = []
        for header in original_headers:
            value = record.original.get(header)
            row.append(prepare_excel_value(value))
        # row = [
        #     record.original.get(header)
        #     for header in original_headers
        # ]

        # Added data
        for _, attribute in OUTPUT_FIELDS.items():

            value = getattr(record, attribute)

            # convert lists for Excel
            if isinstance(value, list):
                value = "; ".join(map(str, value))

            row.append(prepare_excel_value(value))

        ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = "DD.MM.YYYY"

    wb.save(filename)

    return a_counter_is_never_too_much

def data_handler(filename, output_file, type_processing):
    records = read_parse_to_class(filename, type_processing)

    if type_processing=="i":
        data_mask(records)
    elif type_processing=="esg_main":
        process_records(records, type_processing)
    else:
        return "\033[91m Aborted: specify type processing.\033[0m"

    counter = write_records(output_file, records)

    return f"\033[92m{counter} written successfully!\033[0m"



if __name__ == '__main__':
    # xls_parse_base(r"C:\drob\ress.xlsx")
    # xls_parse_sample(r"C:\drob\sample.xlsx")
    # xls_parse_test(r"C:\drob\orf.xlsx")
    #Todo: add checks for brands and types from EAA db.
    #TODO - 2: export keywords []

    # records = read_parse_to_class(r"C:\drob\sample_fordev.xlsx")

    data_handler(r"C:\drob\sample_fordev_u.xlsx",
                 r"C:\drob\upppadted1.xlsx", "i")

    # print(records[0].brand)
    # print(records[0].original)
    # print(records[0].weight)
